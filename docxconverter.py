from docx import Document
import openpyxl
import os

def read_docx(docx_path):
    doc = Document(docx_path)
    data = {'Question': [], 'Option A': [], 'Option B': [], 'Option C': [], 'Option D': [],
            'Ans Opt': [], 'Explanation': [], 'CISA': [],'CISSP': [], 'CCSP': [],
            'SSCP': [], 'CCSK': [], 'CISM': [],}

    current_question = None
    count = 0

    for paragraph in doc.paragraphs:
        if paragraph.text.lower().startswith('question'):
            current_question = paragraph.text.split(':')[-1]
            data['Question'].append(current_question)
        elif paragraph.text.lower().startswith(('a)')):
            data['Option A'].append(paragraph.text[2 :])
        elif paragraph.text.lower().startswith(('b)')):
            data['Option B'].append(paragraph.text[2 :])
        elif paragraph.text.lower().startswith(('c)')):
            data['Option C'].append(paragraph.text[2 :])
        elif paragraph.text.lower().startswith(('d)')):
            data['Option D'].append(paragraph.text[2 :])
        elif paragraph.text.lower().startswith('answer:'):
           data['Ans Opt'].append(paragraph.text.split( ':')[-1])
        elif paragraph.text.lower().startswith('explanation'):
            data['Explanation'].append(paragraph.text.split(':')[-1])
        elif paragraph.text.lower().startswith('cisa'):
            data['CISA'].append(paragraph.text.split('-')[-1])
        elif paragraph.text.lower().startswith('cissp'):
            data['CISSP'].append(paragraph.text.split('-')[-1])
        elif paragraph.text.lower().startswith('ccsp'):
            data['CCSP'].append(paragraph.text.split('-')[-1])
        elif paragraph.text.lower().startswith('sscp'):
            data['SSCP'].append(paragraph.text.split('-')[-1])
        elif paragraph.text.lower().startswith('ccsk'):
            data['CCSK'].append(paragraph.text.split('-')[-1])
        elif paragraph.text.lower().startswith('cism'):
            data['CISM'].append(paragraph.text.split('-')[-1].strip())
       
    # Ensure all lists have the same length
    max_length = max(len(data[key]) for key in data)
    for key in data:
        data[key] += [''] * (max_length - len(data[key]))

    return data

def write_excel(data, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = list(data.keys())
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num in range(2, len(data['Question']) + 2):
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=row_num, column=col_num, value=data[header][row_num - 2])

    wb.save(excel_path)